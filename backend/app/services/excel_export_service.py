"""Service for generating Excel exports of shortlists."""

from datetime import datetime
from io import BytesIO
import json
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.candidate import Candidate
from app.models.job_description import JobDescription
from app.models.shortlist import Shortlist
from app.models.shortlist_candidate import ShortlistCandidate


class ExcelExportService:
    """Service for generating Excel exports of shortlists."""

    def __init__(self, session: Session):
        """Initialize the ExcelExportService with a database session.
        
        Args:
            session: SQLAlchemy database session
        """
        self.session = session

    def generate_shortlist_export(self, shortlist_id: UUID) -> tuple[bytes, str]:
        """Generate a formatted Excel workbook for a shortlist.
        
        Args:
            shortlist_id: UUID of the shortlist to export
            
        Returns:
            Tuple of (Excel file bytes, suggested filename)
            
        Raises:
            ValueError: If shortlist not found or has no candidates
        """
        # Fetch shortlist with related data
        shortlist = self.session.query(Shortlist).filter(Shortlist.id == shortlist_id).first()
        if not shortlist:
            raise ValueError(f"Shortlist {shortlist_id} not found")

        # Fetch JD for naming and info
        jd = self.session.query(JobDescription).filter(JobDescription.id == shortlist.jd_id).first()
        if not jd:
            raise ValueError(f"JobDescription {shortlist.jd_id} not found")

        # Fetch candidate IDs from shortlist
        shortlist_candidates = (
            self.session.query(ShortlistCandidate)
            .filter(ShortlistCandidate.shortlist_id == shortlist_id)
            .all()
        )
        
        if not shortlist_candidates:
            raise ValueError("Shortlist has no candidates")

        candidate_ids = [sc.candidate_id for sc in shortlist_candidates]

        # Fetch candidate data in the order they appear in shortlist
        candidates = self.session.query(Candidate).filter(Candidate.id.in_(candidate_ids)).all()
        
        # Create workbook and add data
        wb = Workbook()
        ws = wb.active
        ws.title = "Shortlist"

        # Add headers
        headers = [
            "Rank",
            "Candidate UUID",
            "Zoho Record ID",
            "Zoho Candidate ID",
            "Candidate Name",
            "Email",
            "Phone",
            "Total Experience (years)",
            "Relevant Experience (years)",
            "Current Company",
            "Skills",
            "Current Location",
            "Preferred Location",
            "Notice Period (days)",
            "Degree",
            "Normalized Degree",
            "Current CTC",
            "Expected CTC",
            "Status",
            "Source",
            "Created At",
            "Updated At",
            "Match Metadata",
            "Raw Payload",
        ]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add candidate data
        for rank, candidate in enumerate(candidates, 1):
            skills = ", ".join(candidate.skills) if candidate.skills else ""
            row = [
                rank,
                str(candidate.id),
                candidate.zoho_record_id or "",
                candidate.zoho_candidate_id or "",
                candidate.full_name or "",
                candidate.email or "",
                candidate.phone or "",
                candidate.total_experience_years or "",
                candidate.relevant_experience_years or "",
                candidate.current_company or "",
                skills,
                candidate.current_location or "",
                candidate.preferred_location or "",
                candidate.notice_period_days or "",
                candidate.degree or "",
                candidate.normalized_degree or "",
                candidate.current_ctc or "",
                candidate.expected_ctc or "",
                candidate.status or "",
                candidate.source or "",
                candidate.created_at.isoformat() if candidate.created_at else "",
                candidate.updated_at.isoformat() if candidate.updated_at else "",
                json.dumps(candidate.match_metadata) if candidate.match_metadata else "",
                json.dumps(candidate.raw_payload) if candidate.raw_payload else "",
            ]
            ws.append(row)

        # Add summary section
        ws.append([])  # Blank row
        ws.append(["Export Summary"])
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"Job Description: {jd.title}"])
        ws.append([f"JD Code: {jd.jd_code}"])
        ws.append([f"Candidate Count: {len(candidates)}"])

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 28
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 24
        ws.column_dimensions["K"].width = 36
        ws.column_dimensions["L"].width = 20
        ws.column_dimensions["M"].width = 20
        ws.column_dimensions["N"].width = 18
        ws.column_dimensions["O"].width = 22
        ws.column_dimensions["P"].width = 22
        ws.column_dimensions["Q"].width = 14
        ws.column_dimensions["R"].width = 14
        ws.column_dimensions["S"].width = 14
        ws.column_dimensions["T"].width = 14
        ws.column_dimensions["U"].width = 30
        ws.column_dimensions["V"].width = 30
        ws.column_dimensions["W"].width = 40
        ws.column_dimensions["X"].width = 48

        # Generate filename
        filename = self._generate_filename(jd.title)

        # Write to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_bytes = output.getvalue()

        return file_bytes, filename

    @staticmethod
    def _generate_filename(jd_title: str) -> str:
        """Generate a descriptive filename from JD title.
        
        Args:
            jd_title: Title of the job description
            
        Returns:
            Formatted filename (e.g., 'Java_Backend_Shortlist.xlsx')
        """
        # Replace spaces with underscores, remove special characters
        sanitized = jd_title.replace(" ", "_").replace("/", "_")
        sanitized = "".join(c for c in sanitized if c.isalnum() or c == "_")
        return f"{sanitized}_Shortlist.xlsx"
